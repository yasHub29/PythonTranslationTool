# modules/excel_translator/excel_writer.py
# Writes translated text back using Excel COM to preserve shapes, text boxes, arrows, and images.
# Falls back to openpyxl if COM fails (for non-Windows or headless environments).

import os
import time
from datetime import datetime

def write_translated_excel_preserve_format(input_path, translations, sheet_renames, output_dir, retry_attempts=3, retry_delay=2):
    """
    Writes translated text back to Excel, preserving formatting when possible (via COM).

    Args:
        input_path (str): Path to the source Excel workbook.
        translations (list): List of tuples (sheet_name, cell_coord, translated_text).
        sheet_renames (list): List of tuples (original_sheet_name, target_sheet_name).
        output_dir (str): Directory to save the translated Excel file.
        retry_attempts (int): Retry count for locked files.
        retry_delay (int): Delay in seconds between retries.

    Returns:
        str: Path to the saved translated Excel file.
    """
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"{base_name}_translated_{timestamp}.xlsx")

    os.makedirs(output_dir, exist_ok=True)

    # --- Try Excel COM first ---
    try:
        import win32com.client as win32
        excel = None
        wb = None

        try:
            excel = win32.gencache.EnsureDispatch("Excel.Application")
        except Exception:
            excel = win32.Dispatch("Excel.Application")  # fallback if cache broken

        excel.DisplayAlerts = False
        excel.Visible = False

        # --- Retry logic for locked files ---
        for attempt in range(retry_attempts):
            try:
                wb = excel.Workbooks.Open(os.path.abspath(input_path), ReadOnly=False)
                break
            except Exception:
                if attempt < retry_attempts - 1:
                    print(f"⚠️ File locked or in use, retrying in {retry_delay}s...")
                    time.sleep(retry_delay)
                else:
                    print("⚠️ File is locked — opening as read-only.")
                    wb = excel.Workbooks.Open(os.path.abspath(input_path), ReadOnly=True)
                    break

        if wb is None:
            raise RuntimeError("Failed to open workbook after multiple attempts.")

        # --- Write translations ---
        for sheet_name, cell_coord, translated_text in translations:
            try:
                ws = wb.Worksheets(sheet_name)

                # Unprotect if necessary
                if ws.ProtectContents:
                    try:
                        ws.Unprotect(Password="")
                    except Exception:
                        pass

                ws.Range(cell_coord).Value = translated_text
            except Exception:
                continue  # skip invalid sheet/cell

        # --- Rename sheets safely ---
        used_names = set()
        for orig_name, target_name in sheet_renames:
            try:
                ws = wb.Worksheets(orig_name)
                if ws.ProtectContents:
                    try:
                        ws.Unprotect(Password="")
                    except Exception:
                        pass

                candidate = target_name
                counter = 1
                while candidate in used_names:
                    candidate = f"{target_name}_{counter}"
                    counter += 1
                used_names.add(candidate)

                ws.Name = candidate
            except Exception:
                continue

        # --- Save as new translated file ---
        wb.SaveAs(os.path.abspath(output_path), FileFormat=51)  # 51 = .xlsx
        print(f"✅ Translated workbook saved at:\n{output_path}")

        # --- Cleanup ---
        wb.Close(SaveChanges=False)
        excel.Quit()

        return output_path

    except Exception as com_exc:
        print("⚠️ Excel COM operation failed — falling back to openpyxl...")
        print(f"Reason: {com_exc}")

        # --- Fallback using openpyxl ---
        try:
            from openpyxl import load_workbook
            wb = load_workbook(input_path)

            # Apply translations
            for sheet_name, cell_coord, translated_text in translations:
                try:
                    ws = wb[sheet_name]
                    ws[cell_coord].value = translated_text
                except Exception:
                    continue

            # Apply sheet renames safely
            existing_names = [s.title for s in wb.worksheets]
            for orig_name, target_name in sheet_renames:
                try:
                    if orig_name in wb.sheetnames:
                        safe_name = target_name[:31] if target_name else target_name
                        final_name = safe_name
                        counter = 1
                        while final_name in existing_names:
                            final_name = f"{safe_name}_{counter}"
                            counter += 1
                        ws = wb[orig_name]
                        ws.title = final_name
                        existing_names.append(final_name)
                except Exception:
                    continue

            wb.save(output_path)
            print(f"✅ Fallback successful: saved with openpyxl at\n{output_path}")
            return output_path

        except Exception as fallback_exc:
            raise RuntimeError(
                f"Both COM and openpyxl operations failed.\n"
                f"COM Error: {com_exc}\nOpenpyxl Error: {fallback_exc}"
            )

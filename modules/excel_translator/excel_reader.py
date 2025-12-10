# modules/excel_translator/excel_reader.py
# Excel file reader - collects text for translation while preserving workbook object

from openpyxl import load_workbook

def read_excel_for_translation(file_path):
    """Reads Excel workbook and collects text from each cell."""
    wb = load_workbook(file_path)
    cells_to_translate = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip():
                    cells_to_translate.append((sheet.title, cell.coordinate, cell.value))
    return wb, cells_to_translate
